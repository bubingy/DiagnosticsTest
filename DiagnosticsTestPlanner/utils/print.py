# coding=utf-8

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

from utils.conf import TestConf


def init_sos_unit_content(os_name: str):
    if 'alpine' in os_name.lower():
        return 'NA'
    else:
        return ''


def init_dump_unit_content(os_name: str):
    if 'osx' in os_name.lower():
        return 'NA'
    else:
        return ''


def init_benchmarks_unit_content(sdk_version: str):
    if sdk_version[0] == '3':
        return ''
    else:
        return 'NA'


def print_weekly_test_matrix(os_rotation: dict, output_file: os.PathLike) -> None:
    workbook = load_workbook(output_file)
    if 'TestMatrix' not in workbook.sheetnames:
        workbook.create_sheet('TestMatrix')
    test_matrix_sheet = workbook['TestMatrix']
    header = [
        'OS',
        'dotnet-counters',
        'dotnet-dump',
        'dotnet-gcdump',
        'dotnet-sos',
        'dotnet-trace',
        'benchmarks'
    ]
    for idx, title in enumerate(header):
        test_matrix_sheet.cell(
            row=1, column=1+idx,
            value=title
        ).font = Font(color="1F497D")

    # write required os
    required_oses = os_rotation['requiredOS']
    current_row = 2
    for idx, key in enumerate(required_oses.keys()):
        os_name = key
        sdk_version = required_oses[key]
        if 'alpine' in key and '6' in required_oses[key]:
            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version}'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            current_row += 1

            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version} disable SYS_PTACE, seccomp=unconfined'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            continue
        
        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version}'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=7,
            value=init_benchmarks_unit_content(sdk_version)
        ).font = Font(color="5B9BD5")
    current_row = current_row + idx + 1

    # write alternate os
    alternate_oses = os_rotation['alternateOS'] 
    for idx, key in enumerate(alternate_oses.keys()):
        os_name = alternate_oses[key]
        sdk_version = key
        if '6' in key:
            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version}'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            current_row += 1

            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version} disable SYS_PTACE, seccomp=unconfined'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            continue

        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version}'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=7,
            value=init_benchmarks_unit_content(sdk_version)
        ).font = Font(color="5B9BD5")

    workbook.save(output_file)


def print_weekly_test_plan(os_rotation, sdk_version, tool_info, output_file) -> None:
    workbook = Workbook()

    # main sheet
    main_sheet = workbook.active
    main_sheet.title = 'period'
    monday = '/'.join(os_rotation['monday'].split('-'))
    friday = '/'.join(os_rotation['friday'].split('-'))
    print(f'Week({monday} ~ {friday})')
    main_sheet.cell(
        row=1, column=1,
        value=f'Week({monday} ~ {friday})'
    )

    # sdk version sheet
    if 'SDKVersion' not in workbook.sheetnames:
        workbook.create_sheet('SDKVersion')
    sdk_version_sheet = workbook['SDKVersion']
    print('Full version of sdk:')
    sdk_version_sheet.cell(
        row=1, column=1,
        value='Full version of sdk:'
    )

    for idx, branch in enumerate(sdk_version.keys()):
        if branch[0] == '3':
            dotnet_prefix = '.net core'
        else:
            dotnet_prefix = '.net'
        print(f'{dotnet_prefix} {branch}: {sdk_version[branch]}')
        sdk_version_sheet.cell(
            row=2+idx, column=1,
            value=f'{dotnet_prefix} {branch}:'
        )
        sdk_version_sheet.cell(
            row=2+idx, column=2,
            value=f'{sdk_version[branch]}'
        )

    # tool info sheet
    if 'ToolInfo' not in workbook.sheetnames:
        workbook.create_sheet('ToolInfo')
    tool_info_sheet = workbook['ToolInfo']
    print('Info of tool:')
    tool_info_sheet.cell(
        row=1, column=1,
        value='Info of tool:'
    )
    print(f'Version: {tool_info.tool_version}')
    tool_info_sheet.cell(
        row=2, column=1,
        value=f'Version:'
    )
    tool_info_sheet.cell(
        row=2, column=2,
        value=f'{tool_info.tool_version}'
    )
    for key in tool_info.pr_info.keys():
        print(f'{key}: {tool_info.pr_info[key]}')
    tool_info_sheet.cell(
        row=3, column=1,
        value='pull'
    )
    tool_info_sheet.cell(
        row=3, column=2,
        value=tool_info.pr_info['pull']
    ).hyperlink = tool_info.pr_info['pull_html_url']
    tool_info_sheet.cell(
        row=3, column=2,
    ).font = Font(color="0563C1")
    tool_info_sheet.cell(
        row=3, column=2,
    ).alignment = Alignment(horizontal='left')
    tool_info_sheet.cell(
        row=4, column=1,
        value='commit'
    )
    tool_info_sheet.cell(
        row=4, column=2,
        value=tool_info.pr_info['commit']
    ).hyperlink = tool_info.pr_info['commit_html_url']
    tool_info_sheet.cell(
        row=4, column=2,
    ).font = Font(color="0563C1")
    workbook.save(output_file)
    print_weekly_test_matrix(os_rotation, output_file)


def print_release_test_matrix(test_conf: TestConf, output_file: os.PathLike) -> None:
    workbook = Workbook()
    main_sheet = workbook.active

    tools = [
        'dotnet-counters',
        'dotnet-dump',
        'dotnet-gcdump',
        'dotnet-sos',
        'dotnet-trace'
    ]

    # print install command 
    main_sheet.title = 'ToolInstallCMD'
    for idx, tool in enumerate(tools): 
        main_sheet.cell(
            row=1+idx, column=1,
            value='dotnet tool install -g {} --version {} --add-source {}'.format(
                tool, test_conf.tool_version, test_conf.tool_feed
            )
        )

    # print test matrix
    header = test_conf.os_list
    for sdk in test_conf.sdk_list:
        sheetname = 'ReleaseTestMatrix-{}'.format(sdk)
        if sheetname not in workbook.sheetnames: workbook.create_sheet(sheetname)
        test_matrix_sheet = workbook[sheetname]
        for idx, title in enumerate(header):
            test_matrix_sheet.cell(
                row=1, column=2+idx,
                value=title
            ).font = Font(bold=True)

        for row_idx, tool in enumerate(tools):
            test_matrix_sheet.cell(
                row=2+row_idx, column=1,
                value=tool
            )
            for col_idx, os in enumerate(header):
                value = ''
                if tool == 'dotnet-dump' and 'osx' in os.lower(): value = 'NA'
                if tool == 'dotnet-sos' and 'alpine' in os.lower(): value = 'NA'
                if value == 'NA':
                    test_matrix_sheet.cell(
                        row=2+row_idx, column=2+col_idx,
                        value=value
                    ).font = Font(color="5B9BD5")

    workbook.save(output_file)
