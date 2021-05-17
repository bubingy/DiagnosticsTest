# coding=utf-8

import os
import json
import base64
import configparser
from urllib import request
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


class TestConf:
    '''Load test configuration.

    '''
    def __init__(self, ini_file_path: os.PathLike) -> None:
        test_conf = configparser.ConfigParser()
        test_conf.read(ini_file_path)
        self.sdk_list = test_conf['SDK']['version'].split('\n')
        self.sdk_list.remove('')
        self.tool_version = test_conf['Tool']['version']
        self.tool_feed = test_conf['Tool']['feed']
        self.os_list = test_conf['Scenarios']['os'].split('\n')
        self.os_list.remove('')


def print_test_matrix(test_conf: TestConf, output_file: os.PathLike) -> None:
    workbook = Workbook()
    main_sheet = workbook.active

    tools = [
        'dotnet-counters',
        'dotnet-dump',
        'dotnet-gcdump',
        'dotnet-sos',
        'dotnet-trace'
    ]

    # print install command 
    main_sheet.title = 'ToolInstallCMD'
    for idx, tool in enumerate(tools): 
        main_sheet.cell(
            row=1+idx, column=1,
            value='dotnet tool install -g {} --version {} --add-source {}'.format(
                tool, test_conf.tool_version, test_conf.tool_feed
            )
        )

    # print test matrix
    header = test_conf.os_list
    for sdk in test_conf.sdk_list:
        sheetname = 'ReleaseTestMatrix-{}'.format(sdk)
        if sheetname not in workbook.sheetnames: workbook.create_sheet(sheetname)
        test_matrix_sheet = workbook[sheetname]
        for idx, title in enumerate(header):
            test_matrix_sheet.cell(
                row=1, column=2+idx,
                value=title
            ).font = Font(bold=True)

        for row_idx, tool in enumerate(tools):
            test_matrix_sheet.cell(
                row=2+row_idx, column=1,
                value=tool
            )
            for col_idx, os in enumerate(header):
                value = ''
                if tool == 'dotnet-dump' and 'osx' in os.lower(): value = 'NA'
                if tool == 'dotnet-sos' and 'alpine' in os.lower(): value = 'NA'
                if value == 'NA':
                    test_matrix_sheet.cell(
                        row=2+row_idx, column=2+col_idx,
                        value=value
                    ).font = Font(color="5B9BD5")

    workbook.save(output_file)


class MQConnectionConf:
    '''Load rabbitmq connection info.

    This class include following properties:
        username: username of rabbitmq.
        password: password.
        ipaddr: ip address of host where rabbitmq run.
        port: port number of rabbitmq-management-plugin.
        vhost: name of vhost.
        base_url: base url for http api.
        general_header: request header for general usage.
    '''
    def __init__(self, ini_file_path: os.PathLike) -> None:
        connection_conf = configparser.ConfigParser()
        connection_conf.read(ini_file_path)
        self.username = connection_conf['connection']['username']
        self.password = connection_conf['connection']['password']
        self.ipaddr = connection_conf['connection']['ipaddr']
        self.port = connection_conf['connection']['port']
        self.vhost = connection_conf['connection']['vhost']
        
        self.base_url = f'http://{self.ipaddr}:{self.port}'
        auth_str = str(
            base64.b64encode(
                bytes(
                    f'{self.username}:{self.password}',
                    'ascii'
                )
            ),
            'ascii'
        )
        self.general_header = {
            'Authorization': f'Basic {auth_str}',
            'content-type':'application/json'
        }


def declare_queue(queue_name: str, connection_conf: MQConnectionConf) -> int:
    '''Declare a queue.

    :param queue_name: name of queue.
    :param connection_conf: rabbitmq connection info.
    :return: 0 if successes, 1 if fails. 
    '''
    data = {
        'auto_delete':'false',
        'durable':'false'
    }
    uri = f'/api/queues/{connection_conf.vhost}/{queue_name}'
    req = request.Request(
        f'{connection_conf.base_url}{uri}',
        headers=connection_conf.general_header,
        data=json.dumps(data).encode("utf-8"),
        method='PUT'
    )
    try:
        request.urlopen(req).read().decode('utf-8')
        return 0
    except Exception as e:
        print(f'fail to declare queue: {e}')
        return 1


def publish_message(message: str, queue_name: str, connection_conf: MQConnectionConf) -> int:
    '''Publish a message.

    :param message: message in string format.
    :param queue_name: name of queue.
    :param connection_conf: rabbitmq connection info.
    :return: 0 if successes, 1 if fails. 
    '''
    data = {
        'properties':{},
        'routing_key':queue_name,
        'payload':message,
        'payload_encoding':'string'
    }
    uri = f'/api/exchanges/{connection_conf.vhost}//publish'
    req = request.Request(
        f'{connection_conf.base_url}{uri}',
        headers=connection_conf.general_header,
        data=json.dumps(data).encode("utf-8"),
        method='POST'
    )
    try:
        request.urlopen(req).read().decode('utf-8')
        return 0
    except Exception as e:
        print(f'fail to publish message: {e}')
        return 1
