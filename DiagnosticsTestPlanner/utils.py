# coding=utf-8

import os
import json
import base64
import configparser
from urllib import request
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment


def load_json(file_path: os.PathLike) -> Any:
    import json
    content = None
    with open(file_path, 'r') as reader:
        content = reader.read()
    content = json.loads(content)
    return content


def init_sos_unit_content(os_name: str):
    if 'alpine' in os_name.lower():
        return 'NA'
    else:
        return ''


def init_dump_unit_content(os_name: str):
    if 'osx' in os_name.lower():
        return 'NA'
    else:
        return ''


def init_benchmarks_unit_content(sdk_version: str):
    if sdk_version[0] == '3':
        return ''
    else:
        return 'NA'


def print_test_matrix(os_rotation, output_file) -> None:
    workbook = load_workbook(output_file)
    if 'TestMatrix' not in workbook.sheetnames:
        workbook.create_sheet('TestMatrix')
    test_matrix_sheet = workbook['TestMatrix']
    header = [
        'OS',
        'dotnet-counters',
        'dotnet-dump',
        'dotnet-gcdump',
        'dotnet-sos',
        'dotnet-trace',
        'benchmarks'
    ]
    for idx, title in enumerate(header):
        test_matrix_sheet.cell(
            row=1, column=1+idx,
            value=title
        ).font = Font(color="1F497D")

    # write required os
    required_oses = os_rotation['requiredOS']
    current_row = 2
    for idx, key in enumerate(required_oses.keys()):
        os_name = key
        sdk_version = required_oses[key]
        if 'alpine' in key and '6' in required_oses[key]:
            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version}'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            current_row += 1

            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version} disable SYS_PTACE, seccomp=unconfined'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            continue
        
        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version}'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=7,
            value=init_benchmarks_unit_content(sdk_version)
        ).font = Font(color="5B9BD5")
    current_row = current_row + idx + 1

    # write alternate os
    alternate_oses = os_rotation['alternateOS'] 
    for idx, key in enumerate(alternate_oses.keys()):
        os_name = alternate_oses[key]
        sdk_version = key
        if '6' in key:
            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version}'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            current_row += 1

            test_matrix_sheet.cell(
                row=current_row+idx, column=1,
                value=f'{os_name}/{sdk_version} disable SYS_PTACE, seccomp=unconfined'
            ).font = Font(color="1F497D")
            test_matrix_sheet.cell(
                row=current_row+idx, column=3,
                value=init_dump_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=5,
                value=init_sos_unit_content(os_name)
            ).font = Font(color="5B9BD5")
            test_matrix_sheet.cell(
                row=current_row+idx, column=7,
                value=init_benchmarks_unit_content(sdk_version)
            ).font = Font(color="5B9BD5")
            continue

        test_matrix_sheet.cell(
            row=current_row+idx, column=1,
            value=f'{os_name}/{sdk_version}'
        ).font = Font(color="1F497D")
        test_matrix_sheet.cell(
            row=current_row+idx, column=3,
            value=init_dump_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=5,
            value=init_sos_unit_content(os_name)
        ).font = Font(color="5B9BD5")
        test_matrix_sheet.cell(
            row=current_row+idx, column=7,
            value=init_benchmarks_unit_content(sdk_version)
        ).font = Font(color="5B9BD5")

    workbook.save(output_file)


def print_test_result_page(os_rotation, sdk_version, tool_info, output_file) -> None:
    workbook = Workbook()

    # main sheet
    main_sheet = workbook.active
    main_sheet.title = 'period'
    monday = '/'.join(os_rotation['monday'].split('-'))
    friday = '/'.join(os_rotation['friday'].split('-'))
    print(f'Week({monday} ~ {friday})')
    main_sheet.cell(
        row=1, column=1,
        value=f'Week({monday} ~ {friday})'
    )

    # sdk version sheet
    if 'SDKVersion' not in workbook.sheetnames:
        workbook.create_sheet('SDKVersion')
    sdk_version_sheet = workbook['SDKVersion']
    print('Full version of sdk:')
    sdk_version_sheet.cell(
        row=1, column=1,
        value='Full version of sdk:'
    )

    for idx, branch in enumerate(sdk_version.keys()):
        if branch[0] == '3':
            dotnet_prefix = '.net core'
        else:
            dotnet_prefix = '.net'
        print(f'{dotnet_prefix} {branch}: {sdk_version[branch]}')
        sdk_version_sheet.cell(
            row=2+idx, column=1,
            value=f'{dotnet_prefix} {branch}:'
        )
        sdk_version_sheet.cell(
            row=2+idx, column=2,
            value=f'{sdk_version[branch]}'
        )

    # tool info sheet
    if 'ToolInfo' not in workbook.sheetnames:
        workbook.create_sheet('ToolInfo')
    tool_info_sheet = workbook['ToolInfo']
    print('Info of tool:')
    tool_info_sheet.cell(
        row=1, column=1,
        value='Info of tool:'
    )
    print(f'Version: {tool_info.tool_version}')
    tool_info_sheet.cell(
        row=2, column=1,
        value=f'Version:'
    )
    tool_info_sheet.cell(
        row=2, column=2,
        value=f'{tool_info.tool_version}'
    )
    for key in tool_info.pr_info.keys():
        print(f'{key}: {tool_info.pr_info[key]}')
    tool_info_sheet.cell(
        row=3, column=1,
        value='pull'
    )
    tool_info_sheet.cell(
        row=3, column=2,
        value=tool_info.pr_info['pull']
    ).hyperlink = tool_info.pr_info['pull_html_url']
    tool_info_sheet.cell(
        row=3, column=2,
    ).font = Font(color="0563C1")
    tool_info_sheet.cell(
        row=3, column=2,
    ).alignment = Alignment(horizontal='left')
    tool_info_sheet.cell(
        row=4, column=1,
        value='commit'
    )
    tool_info_sheet.cell(
        row=4, column=2,
        value=tool_info.pr_info['commit']
    ).hyperlink = tool_info.pr_info['commit_html_url']
    tool_info_sheet.cell(
        row=4, column=2,
    ).font = Font(color="0563C1")
    workbook.save(output_file)
    print_test_matrix(os_rotation, output_file)


class AMQPConnectionConf:
    '''Load rabbitmq connection info.

    This class include following properties:
        username: username of rabbitmq.
        password: password.
        ipaddr: ip address of host where rabbitmq run.
        port: port number of rabbitmq-management-plugin.
        vhost: name of vhost.
        base_url: base url for http api.
        general_header: request header for general usage.
    '''
    def __init__(self, ini_file_path: os.PathLike) -> None:
        connection_conf = configparser.ConfigParser()
        connection_conf.read(ini_file_path)
        self.username = connection_conf['connection']['username']
        self.password = connection_conf['connection']['password']
        self.ipaddr = connection_conf['connection']['ipaddr']
        self.port = connection_conf['connection']['port']
        self.vhost = connection_conf['connection']['vhost']
        
        self.base_url = f'http://{self.ipaddr}:{self.port}'
        auth_str = str(
            base64.b64encode(
                bytes(
                    f'{self.username}:{self.password}',
                    'ascii'
                )
            ),
            'ascii'
        )
        self.general_header = {
            'Authorization': f'Basic {auth_str}',
            'content-type':'application/json'
        }


def declare_queue(queue_name: str, connection_conf: AMQPConnectionConf) -> int:
    '''Declare a queue.

    :param queue_name: name of queue.
    :param connection_conf: rabbitmq connection info.
    :return: 0 if successes, 1 if fails. 
    '''
    data = {
        'auto_delete':'false',
        'durable':'false'
    }
    uri = f'/api/queues/{connection_conf.vhost}/{queue_name}'
    req = request.Request(
        f'{connection_conf.base_url}{uri}',
        headers=connection_conf.general_header,
        data=json.dumps(data).encode("utf-8"),
        method='PUT'
    )
    try:
        request.urlopen(req).read().decode('utf-8')
        return 0
    except Exception as e:
        print(f'fail to declare queue: {e}')
        return 1


def publish_message(message: str, queue_name: str, connection_conf: AMQPConnectionConf) -> int:
    '''Publish a message.

    :param message: message in string format.
    :param queue_name: name of queue.
    :param connection_conf: rabbitmq connection info.
    :return: 0 if successes, 1 if fails. 
    '''
    data = {
        'properties':{},
        'routing_key':queue_name,
        'payload':message,
        'payload_encoding':'string'
    }
    uri = f'/api/exchanges/{connection_conf.vhost}//publish'
    req = request.Request(
        f'{connection_conf.base_url}{uri}',
        headers=connection_conf.general_header,
        data=json.dumps(data).encode("utf-8"),
        method='POST'
    )
    try:
        request.urlopen(req).read().decode('utf-8')
        return 0
    except Exception as e:
        print(f'fail to publish message: {e}')
        return 1
